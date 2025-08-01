import sys
import os
import traci
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

# Set SUMO environment path
os.environ['SUMO_HOME'] = r"C:\Program Files (x86)\Eclipse\Sumo"
tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
sys.path.append(tools)

# SUMO config
sumoBinary = os.path.join(os.environ['SUMO_HOME'], 'bin', 'sumo-gui.exe')
sumoConfig = r"D:\test_sumo_1\Test_2_Meeting.sumocfg"
sumoCmd = [sumoBinary, "-c", sumoConfig]

# Start simulation
traci.start(sumoCmd)
step = 0

# Intersection center
intersection_x = 0
intersection_y = 0

# Store previous speeds for acceleration calculation
prev_speeds = {}

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Vehicle Data"
header = ['Step', 'VehicleID', 'PosX', 'PosY', 'DistanceToIntersection_m', 'Speed_kmh', 'Acceleration_mps2', 'Visibility']
ws.append(header)

# Align header
for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
        cell.alignment = Alignment(horizontal='center')

print("\nTracking vehicle distance, speed, acceleration and visibility...\n")

try:
    while traci.simulation.getMinExpectedNumber() > 0:
        traci.simulationStep()
        step += 1

        vehicle_ids = traci.vehicle.getIDList()
        if not vehicle_ids:
            continue

        for vid in vehicle_ids:
            try:
                x, y = traci.vehicle.getPosition(vid)
                distance = math.sqrt((x - intersection_x) ** 2 + (y - intersection_y) ** 2)
                speed_mps = traci.vehicle.getSpeed(vid)
                speed_kmh = speed_mps * 3.6
                prev_speed = prev_speeds.get(vid, 0.0)
                acceleration = speed_mps - prev_speed
                prev_speeds[vid] = speed_mps

                # Get visibility
                try:
                    visibility = traci.vehicle.getParameter(vid, "junctionModel.visibility")
                except:
                    visibility = "N/A"

                print(f"Step {step}: Vehicle {vid} | Pos: x={x:.2f}, y={y:.2f} | Distance: {distance:.2f} m | Speed: {speed_kmh:.2f} km/h | Accel: {acceleration:.2f} m/s² | Visibility: {visibility}")

                # Write to Excel
                row = [
                    step, vid,
                    round(x, 2), round(y, 2),
                    round(distance, 2),
                    round(speed_kmh, 2),
                    round(acceleration, 2),
                    visibility
                ]
                ws.append(row)

                # Align cells
                current_row = ws.max_row
                ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
                ws[f"B{current_row}"].alignment = Alignment(horizontal='left')
                for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
                    ws[f"{col_letter}{current_row}"].alignment = Alignment(horizontal='right')

            except Exception as e:
                print(f"Error tracking vehicle {vid}: {e}")

finally:
    traci.close()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_path = fr"D:\Test_Meeting\vehicle_tracking_{timestamp}.xlsx"
    wb.save(excel_path)
    print(f"\nSimulation ended. Data written to: {excel_path}")
